"""
Script to fetch real scientific articles from CrossRef API and Europe PMC,
fill an Excel spreadsheet, and download available open-access PDFs.
"""

import requests
import openpyxl
import os
import time
import re
import json
from urllib.parse import quote

EXCEL_PATH = r"e:\Documentos\GitHub\QuimicaDeAltura_LandingPage\articles\ARTICULOS .xlsx"
PDF_DIR = r"e:\Documentos\GitHub\QuimicaDeAltura_LandingPage\articles\pdfs"

# Search queries mapped to classification labels
TOPICS = [
    {
        "query": "plant genomic analysis sequencing",
        "classification": "Plant Genomic Analysis",
        "count": 17,
    },
    {
        "query": "artificial intelligence plant recognition deep learning",
        "classification": "AI Plant Recognition",
        "count": 17,
    },
    {
        "query": "machine learning plant pharmacology prediction medicinal",
        "classification": "Plant Pharmacology Prediction Apps",
        "count": 17,
    },
    {
        "query": "phytochemical analysis natural products plant identification",
        "classification": "Natural Products Analysis",
        "count": 16,
    },
    {
        "query": "technology medicinal plants ethnobotany digital",
        "classification": "Technology & Medicinal Plants",
        "count": 16,
    },
]


def search_crossref(query, rows=20, offset=0):
    """Search CrossRef API for articles."""
    url = "https://api.crossref.org/works"
    params = {
        "query": query,
        "rows": rows,
        "offset": offset,
        "filter": "type:journal-article",
        "select": "DOI,title,author,published-print,published-online,container-title,subject,link,URL,abstract",
        "sort": "relevance",
        "order": "desc",
    }
    headers = {
        "User-Agent": "ArticleFetcher/1.0 (mailto:research@example.com)",
    }
    try:
        resp = requests.get(url, params=params, headers=headers, timeout=30)
        resp.raise_for_status()
        return resp.json().get("message", {}).get("items", [])
    except Exception as e:
        print(f"  CrossRef error: {e}")
        return []


def search_europepmc(query, page_size=20, cursor="*"):
    """Search Europe PMC for open access articles."""
    url = "https://www.ebi.ac.uk/europepmc/webservices/rest/search"
    params = {
        "query": query,
        "format": "json",
        "pageSize": page_size,
        "cursorMark": cursor,
        "resultType": "core",
    }
    try:
        resp = requests.get(url, params=params, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        return data.get("resultList", {}).get("result", []), data.get("nextCursorMark")
    except Exception as e:
        print(f"  EuropePMC error: {e}")
        return [], None


def get_authors_str(authors_list):
    """Format authors from CrossRef format."""
    if not authors_list:
        return ""
    names = []
    for a in authors_list:
        family = a.get("family", "")
        given = a.get("given", "")
        if family and given:
            names.append(f"{family}, {given}")
        elif family:
            names.append(family)
    return "; ".join(names)


def get_year(item):
    """Extract publication year from CrossRef item."""
    for field in ["published-print", "published-online", "created"]:
        if field in item:
            parts = item[field].get("date-parts", [[None]])
            if parts and parts[0] and parts[0][0]:
                return parts[0][0]
    return None


def get_keywords(item):
    """Extract keywords/subjects."""
    subjects = item.get("subject", [])
    if subjects:
        return "; ".join(subjects[:5])
    return ""


def sanitize_filename(name):
    """Make a string safe for use as a filename."""
    name = re.sub(r'[<>:"/\\|?*]', '', name)
    name = name.strip()
    return name[:100]


def try_download_pdf(doi, title, idx):
    """Try to download PDF via Unpaywall API (free, open access)."""
    if not doi:
        return False

    # Try Unpaywall
    try:
        url = f"https://api.unpaywall.org/v2/{doi}?email=research@example.com"
        resp = requests.get(url, timeout=15)
        if resp.status_code == 200:
            data = resp.json()
            best = data.get("best_oa_location")
            if best:
                pdf_url = best.get("url_for_pdf") or best.get("url")
                if pdf_url:
                    return download_file(pdf_url, title, idx)
    except Exception:
        pass

    # Try Europe PMC for full text
    try:
        clean_doi = doi.replace("https://doi.org/", "").replace("http://doi.org/", "")
        url = f"https://www.ebi.ac.uk/europepmc/webservices/rest/search?query=DOI:{clean_doi}&format=json&resultType=core"
        resp = requests.get(url, timeout=15)
        if resp.status_code == 200:
            results = resp.json().get("resultList", {}).get("result", [])
            if results:
                r = results[0]
                if r.get("isOpenAccess") == "Y":
                    pmcid = r.get("pmcid")
                    if pmcid:
                        pdf_url = f"https://europepmc.org/backend/ptpmcrender.fcgi?accid={pmcid}&blobtype=pdf"
                        return download_file(pdf_url, title, idx)
    except Exception:
        pass

    return False


def download_file(url, title, idx):
    """Download a file from URL."""
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        }
        resp = requests.get(url, headers=headers, timeout=30, stream=True, allow_redirects=True)
        content_type = resp.headers.get("Content-Type", "")

        if resp.status_code == 200 and ("pdf" in content_type or url.endswith(".pdf") or len(resp.content) > 10000):
            # Check if it's actually a PDF (starts with %PDF)
            first_bytes = resp.content[:5]
            if first_bytes.startswith(b"%PDF") or "pdf" in content_type:
                safe_title = sanitize_filename(title)
                filename = f"{idx:03d}_{safe_title}.pdf"
                filepath = os.path.join(PDF_DIR, filename)
                with open(filepath, "wb") as f:
                    f.write(resp.content)
                print(f"    PDF downloaded: {filename}")
                return True
    except Exception as e:
        pass
    return False


def main():
    os.makedirs(PDF_DIR, exist_ok=True)

    print("Loading Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # Find empty rows (have number but no title)
    empty_rows = []
    for row_idx in range(2, ws.max_row + 1):
        num = ws.cell(row=row_idx, column=1).value
        title = ws.cell(row=row_idx, column=4).value
        if num is not None and title is None:
            empty_rows.append(row_idx)

    print(f"Found {len(empty_rows)} empty rows to fill")

    if not empty_rows:
        print("No empty rows to fill!")
        return

    # Collect articles from all topics
    all_articles = []

    for topic in TOPICS:
        print(f"\nSearching: {topic['classification']}...")
        needed = topic["count"]
        articles = []

        # Search CrossRef with multiple offsets to get enough unique results
        for offset in range(0, needed * 3, 25):
            if len(articles) >= needed:
                break
            items = search_crossref(topic["query"], rows=25, offset=offset)
            time.sleep(1)  # Be polite to the API

            for item in items:
                if len(articles) >= needed:
                    break

                title = item.get("title", [None])[0]
                if not title:
                    continue

                # Skip duplicates
                if any(a["title"] == title for a in articles):
                    continue
                if any(a["title"] == title for a in all_articles):
                    continue

                doi = item.get("DOI", "")
                authors = get_authors_str(item.get("author", []))
                year = get_year(item)
                journal = ", ".join(item.get("container-title", [])) if item.get("container-title") else ""
                keywords = get_keywords(item)

                articles.append({
                    "title": title,
                    "authors": authors,
                    "year": year,
                    "doi": f"https://doi.org/{doi}" if doi else "",
                    "keywords": keywords,
                    "type": "journalArticle",
                    "database": "CrossRef",
                    "classification": topic["classification"],
                    "journal": journal,
                    "raw_doi": doi,
                })

        print(f"  Found {len(articles)} articles")
        all_articles.extend(articles)

    print(f"\nTotal articles collected: {len(all_articles)}")

    # Fill empty rows
    rows_to_fill = min(len(empty_rows), len(all_articles))
    print(f"Filling {rows_to_fill} rows...")

    pdfs_downloaded = 0

    for i in range(rows_to_fill):
        row_idx = empty_rows[i]
        article = all_articles[i]
        num = int(ws.cell(row=row_idx, column=1).value)

        # Fill columns
        ws.cell(row=row_idx, column=2, value=article["classification"])  # Search classification
        ws.cell(row=row_idx, column=3, value=article["year"])  # Year
        ws.cell(row=row_idx, column=4, value=article["title"])  # Title
        ws.cell(row=row_idx, column=5, value=article["authors"])  # Authors
        ws.cell(row=row_idx, column=6, value=article["keywords"])  # Keywords
        ws.cell(row=row_idx, column=7, value=article["type"])  # Type of work
        ws.cell(row=row_idx, column=8, value=article["database"])  # Database
        ws.cell(row=row_idx, column=9, value=article["doi"])  # URL/DOI

        # Try to download PDF
        print(f"  [{i+1}/{rows_to_fill}] #{num}: {article['title'][:60]}...")
        if article["raw_doi"]:
            if try_download_pdf(article["raw_doi"], article["title"], num):
                pdfs_downloaded += 1
            time.sleep(0.5)

    # Save Excel
    print(f"\nSaving Excel...")
    wb.save(EXCEL_PATH)
    print(f"Done! Filled {rows_to_fill} rows, downloaded {pdfs_downloaded} PDFs")


if __name__ == "__main__":
    import sys
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    main()
