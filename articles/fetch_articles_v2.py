"""
Script to fetch real scientific articles, only keep ones with downloadable PDFs.
Uses CrossRef + Unpaywall + Europe PMC + OpenAlex + Semantic Scholar.
Filters for articles from 2020 onwards.
"""

import requests
import openpyxl
import os
import time
import re
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

EXCEL_PATH = r"e:\Documentos\GitHub\QuimicaDeAltura_LandingPage\articles\ARTICULOS .xlsx"
PDF_DIR = r"e:\Documentos\GitHub\QuimicaDeAltura_LandingPage\articles\pdfs"

MIN_YEAR = 2020

# Expanded topics with multiple related queries per classification
TOPICS = [
    {
        "queries": [
            "plant genome sequencing genomic analysis",
            "plant transcriptomics RNA-seq gene expression",
            "crop genomics whole genome sequencing breeding",
            "plant DNA barcoding molecular phylogenetics",
            "chloroplast genome plastome plant evolution",
        ],
        "classification": "Plant Genomic Analysis",
        "target": 17,
    },
    {
        "queries": [
            "deep learning plant species recognition identification",
            "convolutional neural network plant leaf classification",
            "computer vision plant disease detection imaging",
            "artificial intelligence botanical image recognition",
            "transfer learning plant phenotyping automated",
        ],
        "classification": "AI Plant Recognition",
        "target": 17,
    },
    {
        "queries": [
            "machine learning medicinal plant pharmacological activity prediction",
            "QSAR plant natural compound drug discovery",
            "neural network phytochemical bioactivity prediction",
            "computational drug design plant-derived compounds",
            "random forest molecular docking medicinal plants",
        ],
        "classification": "Plant Pharmacology Prediction Apps",
        "target": 17,
    },
    {
        "queries": [
            "phytochemical screening natural products bioactive compounds plant",
            "HPLC mass spectrometry plant metabolite identification",
            "plant secondary metabolites alkaloids flavonoids analysis",
            "antioxidant antimicrobial plant extract characterization",
            "metabolomics plant natural products profiling",
        ],
        "classification": "Natural Products Analysis",
        "target": 16,
    },
    {
        "queries": [
            "digital technology medicinal plants ethnobotany bioinformatics",
            "GIS remote sensing medicinal plant distribution mapping",
            "blockchain traceability herbal medicine supply chain",
            "mobile application traditional medicine plant identification",
            "database informatics ethnopharmacology medicinal plants",
        ],
        "classification": "Technology & Medicinal Plants",
        "target": 16,
    },
]


def sanitize_filename(name):
    name = re.sub(r'[<>:"/\\|?*]', '', name)
    return name.strip()[:100]


# ---------------------------------------------------------------------------
# Database search functions
# ---------------------------------------------------------------------------

def search_europepmc_oa(query, page_size=25, cursor="*"):
    """Search Europe PMC for open access articles from 2020+."""
    url = "https://www.ebi.ac.uk/europepmc/webservices/rest/search"
    params = {
        "query": f"{query} OPEN_ACCESS:y HAS_FT:y FIRST_PDATE:[2020 TO 2026]",
        "format": "json",
        "pageSize": page_size,
        "cursorMark": cursor,
        "resultType": "core",
    }
    try:
        resp = requests.get(url, params=params, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        return data.get("resultList", {}).get("result", []), data.get("nextCursorMark")
    except Exception as e:
        print(f"  EuropePMC error: {e}")
        return [], None


def search_crossref_oa(query, rows=25, offset=0):
    """Search CrossRef for articles from 2020+."""
    url = "https://api.crossref.org/works"
    params = {
        "query": query,
        "rows": rows,
        "offset": offset,
        "filter": "type:journal-article,from-pub-date:2020",
        "select": "DOI,title,author,published-print,published-online,container-title,subject,link,URL",
        "sort": "relevance",
        "order": "desc",
    }
    headers = {"User-Agent": "ArticleFetcher/1.0 (mailto:research@example.com)"}
    try:
        resp = requests.get(url, params=params, headers=headers, timeout=30)
        resp.raise_for_status()
        return resp.json().get("message", {}).get("items", [])
    except Exception as e:
        print(f"  CrossRef error: {e}")
        return []


def search_openalex(query, page=1, per_page=25):
    """Search OpenAlex for open access articles from 2020+."""
    url = "https://api.openalex.org/works"
    params = {
        "search": query,
        "filter": "is_oa:true,from_publication_date:2020-01-01,type:article",
        "sort": "relevance_score:desc",
        "per_page": per_page,
        "page": page,
        "mailto": "research@example.com",
    }
    try:
        resp = requests.get(url, params=params, timeout=30)
        resp.raise_for_status()
        return resp.json().get("results", [])
    except Exception as e:
        print(f"  OpenAlex error: {e}")
        return []


def search_semantic_scholar(query, offset=0, limit=25):
    """Search Semantic Scholar for articles from 2020+."""
    url = "https://api.semanticscholar.org/graph/v1/paper/search"
    params = {
        "query": query,
        "offset": offset,
        "limit": limit,
        "year": "2020-",
        "openAccessPdf": "",
        "fields": "title,authors,year,externalIds,openAccessPdf,journal,tldr",
    }
    try:
        resp = requests.get(url, params=params, timeout=30)
        resp.raise_for_status()
        return resp.json().get("data", [])
    except Exception as e:
        print(f"  SemanticScholar error: {e}")
        return []


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get_authors_str(authors_list):
    if not authors_list:
        return ""
    names = []
    for a in authors_list:
        family = a.get("family", "")
        given = a.get("given", "")
        if family and given:
            names.append(f"{family}, {given}")
        elif family:
            names.append(family)
    return "; ".join(names)


def get_year(item):
    for field in ["published-print", "published-online", "created"]:
        if field in item:
            parts = item[field].get("date-parts", [[None]])
            if parts and parts[0] and parts[0][0]:
                return parts[0][0]
    return None


def try_download_pdf(doi, pmcid, title, idx, oa_pdf_url=None):
    """Try multiple methods to download PDF. Returns True if successful."""
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
    safe_title = sanitize_filename(title)
    filename = f"{idx}_{safe_title}.pdf"
    filepath = os.path.join(PDF_DIR, filename)

    if os.path.exists(filepath):
        return True

    # Method 1: Direct OA PDF URL (from Semantic Scholar / OpenAlex)
    if oa_pdf_url:
        try:
            resp = requests.get(oa_pdf_url, headers=headers, timeout=30, allow_redirects=True)
            if resp.status_code == 200 and resp.content[:5].startswith(b"%PDF"):
                with open(filepath, "wb") as f:
                    f.write(resp.content)
                return True
        except Exception:
            pass

    # Method 2: Europe PMC direct PDF (most reliable for OA)
    if pmcid:
        try:
            pdf_url = f"https://europepmc.org/backend/ptpmcrender.fcgi?accid={pmcid}&blobtype=pdf"
            resp = requests.get(pdf_url, headers=headers, timeout=30, allow_redirects=True)
            if resp.status_code == 200 and resp.content[:5].startswith(b"%PDF"):
                with open(filepath, "wb") as f:
                    f.write(resp.content)
                return True
        except Exception:
            pass

    # Method 3: Unpaywall
    if doi:
        try:
            clean_doi = doi.replace("https://doi.org/", "").replace("http://doi.org/", "")
            url = f"https://api.unpaywall.org/v2/{clean_doi}?email=research@example.com"
            resp = requests.get(url, timeout=15)
            if resp.status_code == 200:
                data = resp.json()
                for loc in data.get("oa_locations", []):
                    pdf_url = loc.get("url_for_pdf")
                    if pdf_url:
                        try:
                            resp2 = requests.get(pdf_url, headers=headers, timeout=30, allow_redirects=True)
                            if resp2.status_code == 200 and resp2.content[:5].startswith(b"%PDF"):
                                with open(filepath, "wb") as f:
                                    f.write(resp2.content)
                                return True
                        except Exception:
                            continue
        except Exception:
            pass

    return False


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    os.makedirs(PDF_DIR, exist_ok=True)

    print("Loading Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # Find empty rows (have number but no title)
    empty_rows = []
    for row_idx in range(2, ws.max_row + 1):
        num = ws.cell(row=row_idx, column=1).value
        title = ws.cell(row=row_idx, column=4).value
        if num is not None and title is None:
            empty_rows.append(row_idx)

    print(f"Found {len(empty_rows)} empty rows to fill")
    if not empty_rows:
        print("No empty rows!")
        return

    # Clear any previously partially filled rows
    for row_idx in empty_rows:
        for col in range(2, 10):
            ws.cell(row=row_idx, column=col, value=None)

    filled_count = 0
    row_pointer = 0
    seen_titles = set()

    # Collect existing titles to avoid duplicates
    for row_idx in range(2, ws.max_row + 1):
        t = ws.cell(row=row_idx, column=4).value
        if t:
            seen_titles.add(str(t).lower().strip())

    for topic in TOPICS:
        target = topic["target"]
        topic_filled = 0
        print(f"\n{'='*60}")
        print(f"  {topic['classification']} (need {target})")
        print(f"{'='*60}")

        for query in topic["queries"]:
            if topic_filled >= target or row_pointer >= len(empty_rows):
                break

            print(f"\n  Query: {query}")

            # --- Database 1: Europe PMC ---
            print(f"  [Europe PMC]")
            cursor = "*"
            attempts = 0
            while topic_filled < target and attempts < 4 and row_pointer < len(empty_rows):
                attempts += 1
                results, cursor = search_europepmc_oa(query, page_size=25, cursor=cursor)
                time.sleep(1)

                if not results:
                    break

                for r in results:
                    if topic_filled >= target or row_pointer >= len(empty_rows):
                        break

                    title = r.get("title", "").strip()
                    if not title or title.lower() in seen_titles:
                        continue

                    year = r.get("pubYear")
                    if year and int(year) < MIN_YEAR:
                        continue

                    pmcid = r.get("pmcid")
                    doi = r.get("doi", "")
                    row_idx = empty_rows[row_pointer]
                    num = int(ws.cell(row=row_idx, column=1).value)

                    print(f"    Trying: {title[:70]}...")
                    if try_download_pdf(doi, pmcid, title, num):
                        authors_list = r.get("authorString", "")
                        keywords = "; ".join(r.get("keywordList", {}).get("keyword", [])[:5]) if r.get("keywordList") else ""

                        ws.cell(row=row_idx, column=2, value=topic["classification"])
                        ws.cell(row=row_idx, column=3, value=int(year) if year else None)
                        ws.cell(row=row_idx, column=4, value=title)
                        ws.cell(row=row_idx, column=5, value=authors_list)
                        ws.cell(row=row_idx, column=6, value=keywords)
                        ws.cell(row=row_idx, column=7, value="journalArticle")
                        ws.cell(row=row_idx, column=8, value="Europe PMC")
                        ws.cell(row=row_idx, column=9, value=f"https://doi.org/{doi}" if doi else (f"https://europepmc.org/article/PMC/{pmcid}" if pmcid else ""))

                        seen_titles.add(title.lower())
                        topic_filled += 1
                        filled_count += 1
                        row_pointer += 1
                        print(f"      OK #{num} ({topic_filled}/{target})")
                    else:
                        print(f"      No PDF, skipping")

                    time.sleep(0.3)

                if not cursor or cursor == "*":
                    break

            # --- Database 2: OpenAlex ---
            if topic_filled < target and row_pointer < len(empty_rows):
                print(f"  [OpenAlex]")
                for page in range(1, 5):
                    if topic_filled >= target or row_pointer >= len(empty_rows):
                        break
                    results = search_openalex(query, page=page, per_page=25)
                    time.sleep(1)

                    if not results:
                        break

                    for r in results:
                        if topic_filled >= target or row_pointer >= len(empty_rows):
                            break

                        title = r.get("title", "")
                        if not title:
                            continue
                        title = title.strip()
                        if title.lower() in seen_titles:
                            continue

                        year = r.get("publication_year")
                        if year and year < MIN_YEAR:
                            continue

                        doi = r.get("doi", "")
                        if doi and doi.startswith("https://doi.org/"):
                            raw_doi = doi.replace("https://doi.org/", "")
                        else:
                            raw_doi = doi or ""

                        # Get OA PDF URL from OpenAlex
                        oa_pdf_url = None
                        oa = r.get("open_access", {})
                        if oa.get("oa_url"):
                            oa_pdf_url = oa.get("oa_url")
                        # Also check primary location
                        primary = r.get("primary_location", {})
                        if primary and primary.get("pdf_url"):
                            oa_pdf_url = primary.get("pdf_url")

                        row_idx = empty_rows[row_pointer]
                        num = int(ws.cell(row=row_idx, column=1).value)

                        print(f"    Trying: {title[:70]}...")
                        if try_download_pdf(raw_doi, None, title, num, oa_pdf_url=oa_pdf_url):
                            # Extract authors
                            authorships = r.get("authorships", [])
                            authors = "; ".join(
                                a.get("author", {}).get("display_name", "")
                                for a in authorships if a.get("author", {}).get("display_name")
                            )
                            # Extract keywords
                            concepts = r.get("concepts", [])
                            keywords = "; ".join(c.get("display_name", "") for c in concepts[:5])

                            ws.cell(row=row_idx, column=2, value=topic["classification"])
                            ws.cell(row=row_idx, column=3, value=year)
                            ws.cell(row=row_idx, column=4, value=title)
                            ws.cell(row=row_idx, column=5, value=authors)
                            ws.cell(row=row_idx, column=6, value=keywords)
                            ws.cell(row=row_idx, column=7, value="journalArticle")
                            ws.cell(row=row_idx, column=8, value="OpenAlex")
                            ws.cell(row=row_idx, column=9, value=f"https://doi.org/{raw_doi}" if raw_doi else "")

                            seen_titles.add(title.lower())
                            topic_filled += 1
                            filled_count += 1
                            row_pointer += 1
                            print(f"      OK #{num} ({topic_filled}/{target})")
                        else:
                            print(f"      No PDF, skipping")

                        time.sleep(0.3)

            # --- Database 3: Semantic Scholar ---
            if topic_filled < target and row_pointer < len(empty_rows):
                print(f"  [Semantic Scholar]")
                for offset in range(0, 100, 25):
                    if topic_filled >= target or row_pointer >= len(empty_rows):
                        break
                    results = search_semantic_scholar(query, offset=offset, limit=25)
                    time.sleep(3)  # Semantic Scholar has stricter rate limits

                    if not results:
                        break

                    for r in results:
                        if topic_filled >= target or row_pointer >= len(empty_rows):
                            break

                        title = r.get("title", "")
                        if not title:
                            continue
                        title = title.strip()
                        if title.lower() in seen_titles:
                            continue

                        year = r.get("year")
                        if year and year < MIN_YEAR:
                            continue

                        ext_ids = r.get("externalIds", {}) or {}
                        doi = ext_ids.get("DOI", "")
                        pmcid = ext_ids.get("PubMedCentral")
                        if pmcid:
                            pmcid = f"PMC{pmcid}" if not pmcid.startswith("PMC") else pmcid

                        oa_pdf = r.get("openAccessPdf", {})
                        oa_pdf_url = oa_pdf.get("url") if oa_pdf else None

                        row_idx = empty_rows[row_pointer]
                        num = int(ws.cell(row=row_idx, column=1).value)

                        print(f"    Trying: {title[:70]}...")
                        if try_download_pdf(doi, pmcid, title, num, oa_pdf_url=oa_pdf_url):
                            authors = "; ".join(
                                a.get("name", "") for a in r.get("authors", []) if a.get("name")
                            )
                            journal_info = r.get("journal", {})
                            journal_name = journal_info.get("name", "") if journal_info else ""

                            ws.cell(row=row_idx, column=2, value=topic["classification"])
                            ws.cell(row=row_idx, column=3, value=year)
                            ws.cell(row=row_idx, column=4, value=title)
                            ws.cell(row=row_idx, column=5, value=authors)
                            ws.cell(row=row_idx, column=6, value="")
                            ws.cell(row=row_idx, column=7, value="journalArticle")
                            ws.cell(row=row_idx, column=8, value="Semantic Scholar")
                            ws.cell(row=row_idx, column=9, value=f"https://doi.org/{doi}" if doi else "")

                            seen_titles.add(title.lower())
                            topic_filled += 1
                            filled_count += 1
                            row_pointer += 1
                            print(f"      OK #{num} ({topic_filled}/{target})")
                        else:
                            print(f"      No PDF, skipping")

                        time.sleep(0.3)

            # --- Database 4: CrossRef + Unpaywall ---
            if topic_filled < target and row_pointer < len(empty_rows):
                print(f"  [CrossRef + Unpaywall]")
                for offset in range(0, 100, 25):
                    if topic_filled >= target or row_pointer >= len(empty_rows):
                        break
                    items = search_crossref_oa(query, rows=25, offset=offset)
                    time.sleep(1)

                    for item in items:
                        if topic_filled >= target or row_pointer >= len(empty_rows):
                            break

                        title_list = item.get("title", [])
                        if not title_list:
                            continue
                        title = title_list[0].strip()
                        if title.lower() in seen_titles:
                            continue

                        year = get_year(item)
                        if year and year < MIN_YEAR:
                            continue

                        doi = item.get("DOI", "")
                        row_idx = empty_rows[row_pointer]
                        num = int(ws.cell(row=row_idx, column=1).value)

                        print(f"    Trying: {title[:70]}...")
                        if try_download_pdf(doi, None, title, num):
                            authors = get_authors_str(item.get("author", []))
                            subjects = item.get("subject", [])
                            keywords = "; ".join(subjects[:5]) if subjects else ""

                            ws.cell(row=row_idx, column=2, value=topic["classification"])
                            ws.cell(row=row_idx, column=3, value=year)
                            ws.cell(row=row_idx, column=4, value=title)
                            ws.cell(row=row_idx, column=5, value=authors)
                            ws.cell(row=row_idx, column=6, value=keywords)
                            ws.cell(row=row_idx, column=7, value="journalArticle")
                            ws.cell(row=row_idx, column=8, value="CrossRef")
                            ws.cell(row=row_idx, column=9, value=f"https://doi.org/{doi}" if doi else "")

                            seen_titles.add(title.lower())
                            topic_filled += 1
                            filled_count += 1
                            row_pointer += 1
                            print(f"      OK #{num} ({topic_filled}/{target})")
                        else:
                            print(f"      No PDF, skipping")

                        time.sleep(0.3)

        print(f"\n  Topic total: {topic_filled}/{target}")

    # Save
    print(f"\nSaving Excel...")
    wb.save(EXCEL_PATH)
    print(f"\nDone! Filled {filled_count}/{len(empty_rows)} rows (all with PDFs downloaded)")
    print(f"PDFs saved in: {PDF_DIR}")


if __name__ == "__main__":
    main()
